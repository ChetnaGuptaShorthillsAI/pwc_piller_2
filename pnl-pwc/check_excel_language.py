import openpyxl
from detect_translate import Translator
# Replace 'your_excel_file.xlsx' with the path to your Excel file
excel_file = "pnl-pwc/Data Points.xlsx"
translator = Translator()

# Open the Excel file in write mode
workbook = openpyxl.load_workbook(excel_file)
write_workbook = openpyxl.Workbook()
sheetnames = workbook.sheetnames
write_workbook.remove(write_workbook.active)

# Loop through all sheets in the original workbook
for sheet_name in sheetnames:
    sheet = workbook[sheet_name]
    write_sheet = write_workbook.create_sheet(title=sheet_name)
    
    # Loop through all rows and columns in the original sheet
    for row in sheet.iter_rows():
        write_row = []
        for cell in row:
            # Edit the cell value and insert a new value
            
            is_english = translator.detect_english_language(str(cell.value))
            if(is_english==False):
                translated_text = translator.convert_to_english(str(cell.value))
                print(translated_text)
                cell.value=translated_text

            write_row.append(cell.value)
        
        write_sheet.append(write_row)

# Save the modified workbook to a new Excel file
write_excel_file = "modified_excel_file.xlsx"
write_workbook.save(write_excel_file)

# Close both workbooks
workbook.close()
write_workbook.close()

print(f"Modified Excel file saved as {write_excel_file}")
